import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Absolute path to the folder containing images
path = 'C:/Users/Arnava/Downloads/Auto_Att-Opencv/ImageAttendance'

# Ensure the ImageAttendance directory exists
if not os.path.exists(path):
    print(f"Error: Directory '{path}' not found. Please check the path.")
    exit()

images = []
classNames = []
myList = os.listdir(path)

# Check if the directory is empty
if not myList:
    print(f"No images found in '{path}'.")
    exit()

print("Images found:", myList)

# Load images and class names
for cl in myList:
    curImg = cv2.imread(f'{path}/{cl}')
    if curImg is None:
        print(f"Warning: Unable to load image '{cl}'. Skipping.")
        continue
    images.append(curImg)
    classNames.append(os.path.splitext(cl)[0])

print("Class names:", classNames)

# Function to find encodings of the images
def findEncodings(images):
    encodeList = []
    for img in images:
        try:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encode = face_recognition.face_encodings(img)[0]  # Taking the first face encoding
            encodeList.append(encode)
        except IndexError:
            print("Warning: No face found in one of the images. Skipping this image.")
    return encodeList

# Function to mark attendance in Excel
def markAttendance(name):
    file_path = 'C:/Users/Arnava/Downloads/Auto_Att-Opencv/Attendance.xlsx'
    now = datetime.now()
    dateString = now.strftime('%Y-%m-%d')
    timeString = now.strftime('%H:%M:%S')
    
    # Load or create workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Name', 'Date', 'Time'])  # Adding header for new file

    # Check for existing entry
    for row in sheet.iter_rows(values_only=True):
        if row[0] == name and row[1] == dateString:
            print(f"{name} already marked for today.")
            return

    # Append new attendance
    sheet.append([name, dateString, timeString])
    workbook.save(file_path)
    print(f"Marked attendance for {name} at {dateString} {timeString}")

# Get encodings for known faces
encodeListKnown = findEncodings(images)
print('Encoding Complete')

# Open the webcam for real-time face recognition
cap = cv2.VideoCapture(0)

if not cap.isOpened():
    print("Error: Webcam not detected.")
    exit()

distance_threshold = 0.5  # Adjusted threshold for recognizing a face

while True:
    success, img = cap.read()
    
    if not success or img is None:
        print("Failed to capture image. Please check your webcam.")
        continue

    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)  # Resize for efficiency
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        
        if len(faceDis) > 0:
            matchIndex = np.argmin(faceDis)
            confidence_score = 1 - faceDis[matchIndex]  # Calculate confidence
            print(f"Match found with {classNames[matchIndex]} - Distance: {faceDis[matchIndex]}, Confidence: {confidence_score:.2f}")

            if matches[matchIndex] and faceDis[matchIndex] < distance_threshold:
                name = classNames[matchIndex].upper()
                print(f"Recognized: {name} with confidence {confidence_score:.2f}")
                
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

                # Mark attendance in Excel
                markAttendance(name)
            else:
                print("No match found or confidence too low.")
                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                cv2.putText(img, "No Data Available", (x1, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 2)

    cv2.imshow('Webcam', img)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

